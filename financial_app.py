import streamlit as st
import pandas as pd
import numpy as np
import re
from datetime import datetime
import json
import os
import base64
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Function to rerun the app based on Streamlit version
def rerun_app():
    try:
        st.rerun()  # For newer Streamlit versions
    except:
        try:
            st.experimental_rerun()  # For older Streamlit versions
        except:
            st.warning("Please refresh the page to see changes.")

# Page configuration
st.set_page_config(
    page_title="Financial Statements Preparation System",
    page_icon="📊",
    layout="wide"
)

# Initialize session state variables if they don't exist
if 'mapped_accounts' not in st.session_state:
    st.session_state.mapped_accounts = {}
if 'tally_data' not in st.session_state:
    st.session_state.tally_data = None
if 'new_ledgers' not in st.session_state:
    st.session_state.new_ledgers = []
if 'versions' not in st.session_state:
    st.session_state.versions = []
if 'current_version' not in st.session_state:
    st.session_state.current_version = None
if 'financial_statements' not in st.session_state:
    st.session_state.financial_statements = None
if 'excel_template' not in st.session_state:
    st.session_state.excel_template = None
if 'sub_schedule_mapping' not in st.session_state:
    st.session_state.sub_schedule_mapping = {}

# Define hierarchy of financial statement categories
FINANCIAL_STATEMENT_HIERARCHY = {
    "Balance Sheet": {
        "Assets": {
            "Fixed Assets": [
                "Land and Buildings",
                "Plant and Machinery",
                "Furniture and Fixtures",
                "Vehicles",
                "Computer Equipment",
                "Other Fixed Assets"
            ],
            "Investments": [
                "Long-term Investments",
                "Short-term Investments",
                "Investment in Properties",
                "Other Investments"
            ],
            "Current Assets": [
                "Inventories",
                "Sundry Debtors",
                "Cash and Bank Balances",
                "Loans and Advances",
                "Deposits",
                "Other Current Assets"
            ]
        },
        "Liabilities": {
            "Capital Account": [
                "Owner's Capital",
                "Partner's Capital",
                "Drawings",
                "Other Capital Items"
            ],
            "Reserves and Surplus": [
                "General Reserve",
                "Revaluation Reserve",
                "Retained Earnings",
                "Other Reserves"
            ],
            "Long Term Loans": [
                "Secured Loans",
                "Unsecured Loans",
                "Term Loans",
                "Other Long-term Loans"
            ],
            "Current Liabilities": [
                "Sundry Creditors",
                "Outstanding Expenses",
                "Statutory Liabilities",
                "Advances from Customers",
                "Other Current Liabilities"
            ]
        }
    },
    "Profit & Loss": {
        "Income": {
            "Revenue from Operations": [
                "Domestic Sales",
                "Export Sales",
                "Service Income",
                "Other Operational Income"
            ],
            "Other Income": [
                "Interest Income",
                "Dividend Income",
                "Rental Income",
                "Miscellaneous Income"
            ]
        },
        "Expenses": {
            "Cost of Goods Sold": [
                "Raw Material Consumed",
                "Direct Expenses",
                "Purchase of Stock-in-Trade",
                "Changes in Inventories"
            ],
            "Employee Benefits": [
                "Salaries and Wages",
                "Bonus and Incentives",
                "Staff Welfare Expenses",
                "Other Employee Costs"
            ],
            "Finance Cost": [
                "Interest Expenses",
                "Bank Charges",
                "Other Financial Charges"
            ],
            "Depreciation": [
                "Depreciation on Fixed Assets",
                "Amortization",
                "Other Depreciation"
            ],
            "Other Expenses": [
                "Administrative Expenses",
                "Selling and Distribution Expenses",
                "Rent and Utilities",
                "Repairs and Maintenance",
                "Travel and Conveyance",
                "Legal and Professional Fees",
                "Insurance",
                "Miscellaneous Expenses"
            ]
        }
    }
}

# Create flattened mapping options for UI
def create_mapping_options():
    options = ["Select mapping..."]
    
    # First level options for main sections
    for statement, categories in FINANCIAL_STATEMENT_HIERARCHY.items():
        for category_type, category_items in categories.items():
            for category, sub_categories in category_items.items():
                # Format: BS_FixedAssets - Fixed Assets
                prefix = "BS_" if statement == "Balance Sheet" else "PL_"
                code = prefix + "".join(category.split())
                option = f"{code} - {category}"
                options.append(option)
    
    return options

# Create flattened sub-category options for UI
def create_sub_category_options(main_category):
    if not main_category or main_category == "Select mapping...":
        return ["Select sub-category..."]
    
    options = ["Select sub-category..."]
    
    # Extract the category name from the option (after the " - ")
    category_name = main_category.split(" - ")[1] if " - " in main_category else main_category
    
    # Find this category in the hierarchy
    for statement, categories in FINANCIAL_STATEMENT_HIERARCHY.items():
        for category_type, category_items in categories.items():
            if category_name in category_items:
                # Add all sub-categories for this main category
                sub_categories = category_items[category_name]
                for sub in sub_categories:
                    # Create a code for the sub-category
                    prefix = "BS_" if statement == "Balance Sheet" else "PL_"
                    main_code = prefix + "".join(category_name.split())
                    sub_code = "".join(sub.split())
                    option = f"{main_code}_{sub_code} - {sub}"
                    options.append(option)
                return options
    
    return ["Select sub-category..."]

# Function to parse Tally text file using regex patterns (for malformed XML)
def parse_tally_file(content):
    # If content is bytes, decode it
    if isinstance(content, bytes):
        content = content.decode('utf-8', errors='replace')
    
    # Clean up common XML issues
    content = content.replace('&*#13;', '')
    content = content.replace('&#10;', '')
    content = content.replace('&#13;', '')
    content = re.sub(r'>\s*\*', '>', content)  # Remove stray asterisks after closing tags
    
    # Find all account names and balances using regex
    # This pattern looks for name and amount pairs in various Tally formats
    ledgers = []
    
    # Try to find ledger names with DSPDISPNAME pattern
    name_pattern = r'<DSPDISPNAME>(.*?)</DSPDISPNAME>'
    amount_pattern = r'<DSPCLDRAMTA>(.*?)</DSPCLDRAMTA>'
    
    names = re.findall(name_pattern, content)
    amounts = re.findall(amount_pattern, content)
    
    # If we found names and amounts, pair them up
    if names and amounts:
        # Make sure we don't go out of bounds
        for i in range(min(len(names), len(amounts))):
            name = names[i].strip()
            amount_str = amounts[i].strip()
            
            # Convert amount to float, handling negative numbers
            try:
                amount = float(amount_str.replace(',', ''))
            except ValueError:
                amount = 0
            
            ledgers.append({
                'name': name,
                'balance': amount
            })
    
    # If the above didn't work, try alternative patterns
    if not ledgers:
        # Try to find ledger names with NAME attribute
        alt_pattern = r'<\w+ NAME="([^"]+)"[^>]*>.*?<\w+>([\d.-]+)</\w+>'
        matches = re.findall(alt_pattern, content)
        
        for name, amount_str in matches:
            try:
                amount = float(amount_str.replace(',', ''))
            except ValueError:
                amount = 0
                
            ledgers.append({
                'name': name,
                'balance': amount
            })
    
    # If we still don't have ledgers, extract any name-like and amount-like patterns
    if not ledgers:
        # This is a more aggressive approach to find anything that looks like a name-amount pair
        # Extract anything that looks like a name (between tags or quotes)
        potential_names = re.findall(r'<[^>]+>([\w\s&,.]+)</[^>]+>|"([\w\s&,.]+)"', content)
        # Extract anything that looks like an amount
        potential_amounts = re.findall(r'>(-?\d+,?\d*\.?\d*)<', content)
        
        # If we found potential names and amounts, pair them up
        if potential_names and potential_amounts:
            # Make each potential name a flat string
            flat_names = []
            for name_tuple in potential_names:
                # Take the non-empty value from the tuple
                for name in name_tuple:
                    if name.strip():
                        flat_names.append(name.strip())
            
            # Use only names that look like ledger names (not too short, not numbers)
            valid_names = [name for name in flat_names if len(name) > 3 and not name.replace(',', '').replace('.', '').isdigit()]
            
            # Make sure we don't go out of bounds
            for i in range(min(len(valid_names), len(potential_amounts))):
                try:
                    amount = float(potential_amounts[i].replace(',', ''))
                except ValueError:
                    amount = 0
                    
                ledgers.append({
                    'name': valid_names[i],
                    'balance': amount
                })
    
    # If we STILL don't have ledgers, create some sample data
    if not ledgers:
        st.warning("Could not extract ledger information from the file. Using sample data instead.")
        ledgers = [
            {'name': 'Capital Account', 'balance': 1000000},
            {'name': 'Fixed Assets', 'balance': 800000},
            {'name': 'Current Assets', 'balance': 700000},
            {'name': 'Reserves and Surplus', 'balance': 500000},
            {'name': 'Revenue', 'balance': 2000000},
            {'name': 'Expenses', 'balance': 1500000},
        ]
    
    # Try to find Tally version
    version_match = re.search(r'<VERSION>(.*?)</VERSION>', content)
    tally_version = version_match.group(1) if version_match else "Unknown"
    
    return {
        'ledgers': ledgers,
        'tally_version': tally_version,
        'export_date': datetime.now().isoformat()
    }

# Function to identify new ledgers
def identify_new_ledgers(current_ledgers, previous_mappings):
    return [ledger for ledger in current_ledgers if ledger['name'] not in previous_mappings]

# Function to generate structured financial statements based on mappings
def generate_financial_statements():
    timestamp = datetime.now().isoformat()
    
    # Initialize structure based on our hierarchy
    financial_statements = {
        'balance_sheet': {
            'assets': {
                'fixed_assets': {
                    'total': 0,
                    'sub_categories': {}
                },
                'investments': {
                    'total': 0,
                    'sub_categories': {}
                },
                'current_assets': {
                    'total': 0,
                    'sub_categories': {}
                }
            },
            'liabilities': {
                'capital': {
                    'total': 0,
                    'sub_categories': {}
                },
                'reserves': {
                    'total': 0,
                    'sub_categories': {}
                },
                'long_term_loans': {
                    'total': 0,
                    'sub_categories': {}
                },
                'current_liabilities': {
                    'total': 0,
                    'sub_categories': {}
                }
            }
        },
        'profit_and_loss': {
            'income': {
                'revenue': {
                    'total': 0,
                    'sub_categories': {}
                },
                'other_income': {
                    'total': 0,
                    'sub_categories': {}
                }
            },
            'expenses': {
                'cogs': {
                    'total': 0,
                    'sub_categories': {}
                },
                'employee_benefits': {
                    'total': 0,
                    'sub_categories': {}
                },
                'finance_costs': {
                    'total': 0,
                    'sub_categories': {}
                },
                'depreciation': {
                    'total': 0,
                    'sub_categories': {}
                },
                'other_expenses': {
                    'total': 0,
                    'sub_categories': {}
                }
            }
        },
        'sub_schedules': {},
        'generated_at': timestamp
    }
    
    # Initialize sub-schedules with the hierarchy structure
    for statement, categories in FINANCIAL_STATEMENT_HIERARCHY.items():
        for category_type, category_items in categories.items():
            for category, sub_categories in category_items.items():
                # Create a key for the sub-schedule
                prefix = "BS_" if statement == "Balance Sheet" else "PL_"
                category_key = prefix + "".join(category.split())
                
                financial_statements['sub_schedules'][category_key] = {
                    'name': category,
                    'items': {}
                }
                
                # Initialize each sub-category
                for sub_category in sub_categories:
                    sub_key = "".join(sub_category.split())
                    financial_statements['sub_schedules'][category_key]['items'][sub_key] = {
                        'name': sub_category,
                        'amount': 0,
                        'ledgers': []
                    }
    
    # Process each ledger and update the financial statement structure
    if st.session_state.tally_data:
        for ledger in st.session_state.tally_data['ledgers']:
            ledger_name = ledger['name']
            balance = ledger['balance']
            
            # Skip if this ledger isn't mapped
            if ledger_name not in st.session_state.mapped_accounts:
                continue
                
            main_mapping = st.session_state.mapped_accounts[ledger_name]
            
            # Skip if no valid mapping
            if main_mapping == "Select mapping...":
                continue
                
            # Get the sub-category mapping if available
            sub_mapping = "Select sub-category..."
            if ledger_name in st.session_state.sub_schedule_mapping:
                sub_mapping = st.session_state.sub_schedule_mapping[ledger_name]
            
            # Extract category code from mapping (e.g., "BS_FixedAssets" from "BS_FixedAssets - Fixed Assets")
            category_code = main_mapping.split(" - ")[0]
            
            # Update the appropriate category based on mapping
            if category_code.startswith('BS_'):
                # Balance Sheet items
                if 'FixedAssets' in category_code:
                    financial_statements['balance_sheet']['assets']['fixed_assets']['total'] += balance
                    category_key = 'BS_FixedAssets'
                elif 'Investments' in category_code:
                    financial_statements['balance_sheet']['assets']['investments']['total'] += balance
                    category_key = 'BS_Investments'
                elif 'CurrentAssets' in category_code:
                    financial_statements['balance_sheet']['assets']['current_assets']['total'] += balance
                    category_key = 'BS_CurrentAssets'
                elif 'Capital' in category_code:
                    financial_statements['balance_sheet']['liabilities']['capital']['total'] += balance
                    category_key = 'BS_CapitalAccount'
                elif 'Reserves' in category_code:
                    financial_statements['balance_sheet']['liabilities']['reserves']['total'] += balance
                    category_key = 'BS_ReservesandSurplus'
                elif 'LongTermLoans' in category_code:
                    financial_statements['balance_sheet']['liabilities']['long_term_loans']['total'] += balance
                    category_key = 'BS_LongTermLoans'
                elif 'CurrentLiabilities' in category_code:
                    financial_statements['balance_sheet']['liabilities']['current_liabilities']['total'] += balance
                    category_key = 'BS_CurrentLiabilities'
            elif category_code.startswith('PL_'):
                # Profit & Loss items
                if 'Revenue' in category_code:
                    financial_statements['profit_and_loss']['income']['revenue']['total'] += balance
                    category_key = 'PL_RevenuefromOperations'
                elif 'OtherIncome' in category_code:
                    financial_statements['profit_and_loss']['income']['other_income']['total'] += balance
                    category_key = 'PL_OtherIncome'
                elif 'COGS' in category_code:
                    financial_statements['profit_and_loss']['expenses']['cogs']['total'] += balance
                    category_key = 'PL_CostofGoodsSold'
                elif 'EmployeeBenefits' in category_code:
                    financial_statements['profit_and_loss']['expenses']['employee_benefits']['total'] += balance
                    category_key = 'PL_EmployeeBenefits'
                elif 'FinanceCost' in category_code:
                    financial_statements['profit_and_loss']['expenses']['finance_costs']['total'] += balance
                    category_key = 'PL_FinanceCost'
                elif 'Depreciation' in category_code:
                    financial_statements['profit_and_loss']['expenses']['depreciation']['total'] += balance
                    category_key = 'PL_Depreciation'
                elif 'OtherExpenses' in category_code:
                    financial_statements['profit_and_loss']['expenses']['other_expenses']['total'] += balance
                    category_key = 'PL_OtherExpenses'
            
            # Update sub-category if selected
            if sub_mapping != "Select sub-category..." and category_key in financial_statements['sub_schedules']:
                # Extract sub-category code
                sub_code = sub_mapping.split(" - ")[0]
                sub_category = sub_code.replace(f"{category_key}_", "")
                
                # Update the sub-category
                if sub_category in financial_statements['sub_schedules'][category_key]['items']:
                    financial_statements['sub_schedules'][category_key]['items'][sub_category]['amount'] += balance
                    financial_statements['sub_schedules'][category_key]['items'][sub_category]['ledgers'].append({
                        'name': ledger_name,
                        'balance': balance
                    })
                else:
                    # If sub-category not found, create it
                    sub_name = sub_mapping.split(" - ")[1] if " - " in sub_mapping else sub_category
                    financial_statements['sub_schedules'][category_key]['items'][sub_category] = {
                        'name': sub_name,
                        'amount': balance,
                        'ledgers': [{
                            'name': ledger_name,
                            'balance': balance
                        }]
                    }
    
    # If no data was processed, use sample data for demonstration
    if not st.session_state.mapped_accounts:
        # Sample balance sheet data
        financial_statements['balance_sheet'] = {
            'assets': {
                'fixed_assets': {
                    'total': 800000,
                    'sub_categories': {
                        'land_and_buildings': 400000,
                        'plant_and_machinery': 200000,
                        'other_fixed_assets': 200000
                    }
                },
                'investments': {
                    'total': 300000,
                    'sub_categories': {
                        'long_term_investments': 200000,
                        'short_term_investments': 100000
                    }
                },
                'current_assets': {
                    'total': 700000,
                    'sub_categories': {
                        'inventory': 300000,
                        'cash_and_bank': 400000
                    }
                }
            },
            'liabilities': {
                'capital': {
                    'total': 1000000,
                    'sub_categories': {
                        'owner_capital': 1000000
                    }
                },
                'reserves': {
                    'total': 500000,
                    'sub_categories': {
                        'general_reserve': 500000
                    }
                },
                'long_term_loans': {
                    'total': 200000,
                    'sub_categories': {
                        'secured_loans': 200000
                    }
                },
                'current_liabilities': {
                    'total': 100000,
                    'sub_categories': {
                        'sundry_creditors': 100000
                    }
                }
            }
        }
        
        # Sample profit & loss data
        financial_statements['profit_and_loss'] = {
            'income': {
                'revenue': {
                    'total': 2000000,
                    'sub_categories': {
                        'domestic_sales': 1500000,
                        'export_sales': 500000
                    }
                },
                'other_income': {
                    'total': 100000,
                    'sub_categories': {
                        'interest_income': 50000,
                        'misc_income': 50000
                    }
                }
            },
            'expenses': {
                'cogs': {
                    'total': 1000000,
                    'sub_categories': {
                        'raw_materials': 800000,
                        'direct_expenses': 200000
                    }
                },
                'employee_benefits': {
                    'total': 500000,
                    'sub_categories': {
                        'salaries': 400000,
                        'staff_welfare': 100000
                    }
                },
                'finance_costs': {
                    'total': 100000,
                    'sub_categories': {
                        'interest_expense': 80000,
                        'bank_charges': 20000
                    }
                },
                'depreciation': {
                    'total': 200000,
                    'sub_categories': {
                        'depreciation_fixed_assets': 200000
                    }
                },
                'other_expenses': {
                    'total': 300000,
                    'sub_categories': {
                        'rent': 100000,
                        'utilities': 100000,
                        'misc_expenses': 100000
                    }
                }
            }
        }
        
        # Populate sub-schedules with sample data
        for category_key, category_data in financial_statements['sub_schedules'].items():
            if category_key == 'BS_FixedAssets':
                category_data['items']['LandandBuildings']['amount'] = 400000
                category_data['items']['PlantandMachinery']['amount'] = 200000
                category_data['items']['OtherFixedAssets']['amount'] = 200000
            elif category_key == 'BS_Investments':
                category_data['items']['LongtermInvestments']['amount'] = 200000
                category_data['items']['ShorttermInvestments']['amount'] = 100000
            # ... and so on for other categories
    
    # Create notes section
    notes = {
        'note1_capital': {
            'opening_balance': financial_statements['balance_sheet']['liabilities']['capital']['total'] * 0.8,
            'additions': financial_statements['balance_sheet']['liabilities']['capital']['total'] * 0.2,
        }
    }
    
    financial_statements['notes'] = notes
    
    # Save current version
    new_version = {
        'id': len(st.session_state.versions) + 1,
        'timestamp': timestamp,
        'mapped_accounts': st.session_state.mapped_accounts.copy(),
        'sub_schedule_mapping': st.session_state.sub_schedule_mapping.copy(),
        'tally_data': st.session_state.tally_data,
        'financial_statements': financial_statements
    }
    
    st.session_state.versions.append(new_version)
    st.session_state.current_version = new_version['id']
    st.session_state.financial_statements = financial_statements

# Function to load saved mappings
def load_mappings():
    mapping_data = {}
    sub_mapping_data = {}
    
    try:
        with open('account_mappings.json', 'r') as f:
            mapping_data = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        mapping_data = {}
    
    try:
        with open('sub_schedule_mappings.json', 'r') as f:
            sub_mapping_data = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        sub_mapping_data = {}
    
    return mapping_data, sub_mapping_data

# Function to save mappings
def save_mappings():
    with open('account_mappings.json', 'w') as f:
        json.dump(st.session_state.mapped_accounts, f)
    
    with open('sub_schedule_mappings.json', 'w') as f:
        json.dump(st.session_state.sub_schedule_mapping, f)
    
    st.success('Mappings saved successfully!')

# Function to apply style to Excel cell
def apply_cell_style(cell, is_header=False, is_total=False, is_subtotal=False, indent_level=0):
    # Set font
    if is_header:
        cell.font = Font(bold=True, size=12)
    elif is_total:
        cell.font = Font(bold=True, size=11)
    elif is_subtotal:
        cell.font = Font(bold=True, size=10)
    else:
        cell.font = Font(size=10)
    
    # Set alignment
    if indent_level > 0:
        cell.alignment = Alignment(horizontal='left', indent=indent_level)
    else:
        cell.alignment = Alignment(horizontal='left')
    
    # Set border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border
    
    # Set background for headers
    if is_header:
        cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

# Function to export to Excel with sub-schedules
def export_to_excel():
    output = BytesIO()
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Balance Sheet
    ws = wb.active
    ws.title = "Balance Sheet"
    
    # Apply formatting to headers
    ws['A1'] = "Financial Statements"
    ws['A1'].font = Font(bold=True, size=14)
    
    ws['A2'] = "Balance Sheet as at " + datetime.now().strftime("%d-%m-%Y")
    ws['A2'].font = Font(bold=True, size=12)
    
    ws['A4'] = "Particulars"
    ws['B4'] = "Note No."
    ws['C4'] = "Amount (₹)"
    
    apply_cell_style(ws['A4'], is_header=True)
    apply_cell_style(ws['B4'], is_header=True)
    apply_cell_style(ws['C4'], is_header=True)
    
    # Add data (from mapped accounts and calculated values)
    row = 5
    
    # EQUITY AND LIABILITIES
    ws[f'A{row}'] = "EQUITY AND LIABILITIES"
    apply_cell_style(ws[f'A{row}'], is_header=True)
    row += 1
    
    # Capital Account
    ws[f'A{row}'] = "Capital Account"
    ws[f'B{row}'] = "1"
    ws[f'C{row}'] = st.session_state.financial_statements['balance_sheet']['liabilities']['capital']['total']
    
    apply_cell_style(ws[f'A{row}'], indent_level=1)
    apply_cell_style(ws[f'B{row}'])
    apply_cell_style(ws[f'C{row}'])
    row += 1
    
    # Reserves and Surplus
    ws[f'A{row}'] = "Reserves and Surplus"
    ws[f'B{row}'] = "2"
    ws[f'C{row}'] = st.session_state.financial_statements['balance_sheet']['liabilities']['reserves']['total']
    
    apply_cell_style(ws[f'A{row}'], indent_level=1)
    apply_cell_style(ws[f'B{row}'])
    apply_cell_style(ws[f'C{row}'])
    row += 1
    
    # Long Term Loans
    ws[f'A{row}'] = "Long Term Loans"
    ws[f'B{row}'] = "3"
    ws[f'C{row}'] = st.session_state.financial_statements['balance_sheet']['liabilities']['long_term_loans']['total']
    
    apply_cell_style(ws[f'A{row}'], indent_level=1)
    apply_cell_style(ws[f'B{row}'])
    apply_cell_style(ws[f'C{row}'])
    row += 1
    
    # Current Liabilities
    ws[f'A{row}'] = "Current Liabilities"
    ws[f'B{row}'] = "4"
    ws[f'C{row}'] = st.session_state.financial_statements['balance_sheet']['liabilities']['current_liabilities']['total']
    
    apply_cell_style(ws[f'A{row}'], indent_level=1)
    apply_cell_style(ws[f'B{row}'])
    apply_cell_style(ws[f'C{row}'])
    row += 1
    
    # Total Liabilities
    total_liabilities = (
        st.session_state.financial_statements['balance_sheet']['liabilities']['capital']['total'] +
        st.session_state.financial_statements['balance_sheet']['liabilities']['reserves']['total'] +
        st.session_state.financial_statements['balance_sheet']['liabilities']['long_term_loans']['total'] +
        st.session_state.financial_statements['balance_sheet']['liabilities']['current_liabilities']['total']
    )
    
    ws[f'A{row}'] = "Total Liabilities"
    ws[f'C{row}'] = total_liabilities
    
    apply_cell_style(ws[f'A{row}'], is_total=True)
    apply_cell_style(ws[f'B{row}'])
    apply_cell_style(ws[f'C{row}'], is_total=True)
    row += 2
    
    # ASSETS
    ws[f'A{row}'] = "ASSETS"
    apply_cell_style(ws[f'A{row}'], is_header=True)
    row += 1
    
    # Fixed Assets
    ws[f'A{row}'] = "Fixed Assets"
    ws[f'B{row}'] = "5"
    ws[f'C{row}'] = st.session_state.financial_statements['balance_sheet']['assets']['fixed_assets']['total']
    
    apply_cell_style(ws[f'A{row}'], indent_level=1)
    apply_cell_style(ws[f'B{row}'])
    apply_cell_style(ws[f'C{row}'])
    row += 1
    
    # Investments
    ws[f'A{row}'] = "Investments"
    ws[f'B{row}'] = "6"
    ws[f'C{row}'] = st.session_state.financial_statements['balance_sheet']['assets']['investments']['total']
    
    apply_cell_style(ws[f'A{row}'], indent_level=1)
    apply_cell_style(ws[f'B{row}'])
    apply_cell_style(ws[f'C{row}'])
    row += 1
    
    # Current Assets
    ws[f'A{row}'] = "Current Assets"
    ws[f'B{row}'] = "7"
    ws[f'C{row}'] = st.session_state.financial_statements['balance_sheet']['assets']['current_assets']['total']
    
    apply_cell_style(ws[f'A{row}'], indent_level=1)
    apply_cell_style(ws[f'B{row}'])
    apply_cell_style(ws[f'C{row}'])
    row += 1
    
    # Total Assets
    total_assets = (
        st.session_state.financial_statements['balance_sheet']['assets']['fixed_assets']['total'] +
        st.session_state.financial_statements['balance_sheet']['assets']['investments']['total'] +
        st.session_state.financial_statements['balance_sheet']['assets']['current_assets']['total']
    )
    
    ws[f'A{row}'] = "Total Assets"
    ws[f'C{row}'] = total_assets
    
    apply_cell_style(ws[f'A{row}'], is_total=True)
    apply_cell_style(ws[f'B{row}'])
    apply_cell_style(ws[f'C{row}'], is_total=True)
    
    # P&L Sheet
    ws2 = wb.create_sheet("Profit and Loss")
    
    # Add headers
    ws2['A1'] = "Financial Statements"
    ws2['A1'].font = Font(bold=True, size=14)
    
    ws2['A2'] = "Statement of Profit and Loss for the year ended " + datetime.now().strftime("%d-%m-%Y")
    ws2['A2'].font = Font(bold=True, size=12)
    
    ws2['A4'] = "Particulars"
    ws2['B4'] = "Note No."
    ws2['C4'] = "Amount (₹)"
    
    apply_cell_style(ws2['A4'], is_header=True)
    apply_cell_style(ws2['B4'], is_header=True)
    apply_cell_style(ws2['C4'], is_header=True)
    
    # Add P&L data
    row = 5
    
    # INCOME
    ws2[f'A{row}'] = "INCOME"
    apply_cell_style(ws2[f'A{row}'], is_header=True)
    row += 1
    
    # Revenue
    ws2[f'A{row}'] = "Revenue from Operations"
    ws2[f'B{row}'] = "8"
    ws2[f'C{row}'] = st.session_state.financial_statements['profit_and_loss']['income']['revenue']['total']
    
    apply_cell_style(ws2[f'A{row}'], indent_level=1)
    apply_cell_style(ws2[f'B{row}'])
    apply_cell_style(ws2[f'C{row}'])
    row += 1
    
    # Other Income
    ws2[f'A{row}'] = "Other Income"
    ws2[f'B{row}'] = "9"
    ws2[f'C{row}'] = st.session_state.financial_statements['profit_and_loss']['income']['other_income']['total']
    
    apply_cell_style(ws2[f'A{row}'], indent_level=1)
    apply_cell_style(ws2[f'B{row}'])
    apply_cell_style(ws2[f'C{row}'])
    row += 1
    
    # Total Income
    total_income = (
        st.session_state.financial_statements['profit_and_loss']['income']['revenue']['total'] +
        st.session_state.financial_statements['profit_and_loss']['income']['other_income']['total']
    )
    
    ws2[f'A{row}'] = "Total Income"
    ws2[f'C{row}'] = total_income
    
    apply_cell_style(ws2[f'A{row}'], is_subtotal=True)
    apply_cell_style(ws2[f'B{row}'])
    apply_cell_style(ws2[f'C{row}'], is_subtotal=True)
    row += 2
    
    # EXPENSES
    ws2[f'A{row}'] = "EXPENSES"
    apply_cell_style(ws2[f'A{row}'], is_header=True)
    row += 1
    
    # COGS
    ws2[f'A{row}'] = "Cost of Goods Sold"
    ws2[f'B{row}'] = "10"
    ws2[f'C{row}'] = st.session_state.financial_statements['profit_and_loss']['expenses']['cogs']['total']
    
    apply_cell_style(ws2[f'A{row}'], indent_level=1)
    apply_cell_style(ws2[f'B{row}'])
    apply_cell_style(ws2[f'C{row}'])
    row += 1
    
    # Employee Benefits
    ws2[f'A{row}'] = "Employee Benefits Expense"
    ws2[f'B{row}'] = "11"
    ws2[f'C{row}'] = st.session_state.financial_statements['profit_and_loss']['expenses']['employee_benefits']['total']
    
    apply_cell_style(ws2[f'A{row}'], indent_level=1)
    apply_cell_style(ws2[f'B{row}'])
    apply_cell_style(ws2[f'C{row}'])
    row += 1
    
    # Finance Costs
    ws2[f'A{row}'] = "Finance Costs"
    ws2[f'B{row}'] = "12"
    ws2[f'C{row}'] = st.session_state.financial_statements['profit_and_loss']['expenses']['finance_costs']['total']
    
    apply_cell_style(ws2[f'A{row}'], indent_level=1)
    apply_cell_style(ws2[f'B{row}'])
    apply_cell_style(ws2[f'C{row}'])
    row += 1
    
    # Depreciation
    ws2[f'A{row}'] = "Depreciation"
    ws2[f'B{row}'] = "13"
    ws2[f'C{row}'] = st.session_state.financial_statements['profit_and_loss']['expenses']['depreciation']['total']
    
    apply_cell_style(ws2[f'A{row}'], indent_level=1)
    apply_cell_style(ws2[f'B{row}'])
    apply_cell_style(ws2[f'C{row}'])
    row += 1
    
    # Other Expenses
    ws2[f'A{row}'] = "Other Expenses"
    ws2[f'B{row}'] = "14"
    ws2[f'C{row}'] = st.session_state.financial_statements['profit_and_loss']['expenses']['other_expenses']['total']
    
    apply_cell_style(ws2[f'A{row}'], indent_level=1)
    apply_cell_style(ws2[f'B{row}'])
    apply_cell_style(ws2[f'C{row}'])
    row += 1
    
    # Total Expenses
    total_expenses = (
        st.session_state.financial_statements['profit_and_loss']['expenses']['cogs']['total'] +
        st.session_state.financial_statements['profit_and_loss']['expenses']['employee_benefits']['total'] +
        st.session_state.financial_statements['profit_and_loss']['expenses']['finance_costs']['total'] +
        st.session_state.financial_statements['profit_and_loss']['expenses']['depreciation']['total'] +
        st.session_state.financial_statements['profit_and_loss']['expenses']['other_expenses']['total']
    )
    
    ws2[f'A{row}'] = "Total Expenses"
    ws2[f'C{row}'] = total_expenses
    
    apply_cell_style(ws2[f'A{row}'], is_subtotal=True)
    apply_cell_style(ws2[f'B{row}'])
    apply_cell_style(ws2[f'C{row}'], is_subtotal=True)
    row += 2
    
    # Profit Before Tax
    profit = total_income - total_expenses
    ws2[f'A{row}'] = "Profit Before Tax"
    ws2[f'C{row}'] = profit
    
    apply_cell_style(ws2[f'A{row}'], is_total=True)
    apply_cell_style(ws2[f'B{row}'])
    apply_cell_style(ws2[f'C{row}'], is_total=True)
    
    # Create Sub-schedule sheets
    # BS Sub-Schedules
    schedule_sheet = wb.create_sheet("BS Schedules")
    
    schedule_sheet['A1'] = "Balance Sheet Schedules"
    schedule_sheet['A1'].font = Font(bold=True, size=14)
    
    current_row = 3
    schedule_num = 1
    
    # Add each schedule from our financial statements
    for category_key, category_data in st.session_state.financial_statements['sub_schedules'].items():
        # Only process Balance Sheet schedules here
        if not category_key.startswith('BS_'):
            continue
            
        # Get category name
        category_name = category_data['name']
        
        # Add schedule header
        schedule_sheet[f'A{current_row}'] = f"Schedule {schedule_num}: {category_name}"
        schedule_sheet[f'A{current_row}'].font = Font(bold=True, size=12)
        current_row += 2
        
        # Add column headers
        schedule_sheet[f'A{current_row}'] = "Particulars"
        schedule_sheet[f'B{current_row}'] = "Amount (₹)"
        
        apply_cell_style(schedule_sheet[f'A{current_row}'], is_header=True)
        apply_cell_style(schedule_sheet[f'B{current_row}'], is_header=True)
        current_row += 1
        
        # Add sub-category items
        total = 0
        for sub_key, sub_data in category_data['items'].items():
            amount = sub_data['amount']
            total += amount
            
            schedule_sheet[f'A{current_row}'] = sub_data['name']
            schedule_sheet[f'B{current_row}'] = amount
            
            apply_cell_style(schedule_sheet[f'A{current_row}'], indent_level=1)
            apply_cell_style(schedule_sheet[f'B{current_row}'])
            current_row += 1
        
        # Add total
        schedule_sheet[f'A{current_row}'] = f"Total {category_name}"
        schedule_sheet[f'B{current_row}'] = total
        
        apply_cell_style(schedule_sheet[f'A{current_row}'], is_total=True)
        apply_cell_style(schedule_sheet[f'B{current_row}'], is_total=True)
        current_row += 3
        
        schedule_num += 1
    
    # PL Sub-Schedules
    pl_schedule_sheet = wb.create_sheet("PL Schedules")
    
    pl_schedule_sheet['A1'] = "Profit & Loss Schedules"
    pl_schedule_sheet['A1'].font = Font(bold=True, size=14)
    
    current_row = 3
    schedule_num = 8  # Continue schedule numbering from BS
    
    # Add each schedule from our financial statements
    for category_key, category_data in st.session_state.financial_statements['sub_schedules'].items():
        # Only process P&L schedules here
        if not category_key.startswith('PL_'):
            continue
            
        # Get category name
        category_name = category_data['name']
        
        # Add schedule header
        pl_schedule_sheet[f'A{current_row}'] = f"Schedule {schedule_num}: {category_name}"
        pl_schedule_sheet[f'A{current_row}'].font = Font(bold=True, size=12)
        current_row += 2
        
        # Add column headers
        pl_schedule_sheet[f'A{current_row}'] = "Particulars"
        pl_schedule_sheet[f'B{current_row}'] = "Amount (₹)"
        
        apply_cell_style(pl_schedule_sheet[f'A{current_row}'], is_header=True)
        apply_cell_style(pl_schedule_sheet[f'B{current_row}'], is_header=True)
        current_row += 1
        
        # Add sub-category items
        total = 0
        for sub_key, sub_data in category_data['items'].items():
            amount = sub_data['amount']
            total += amount
            
            pl_schedule_sheet[f'A{current_row}'] = sub_data['name']
            pl_schedule_sheet[f'B{current_row}'] = amount
            
            apply_cell_style(pl_schedule_sheet[f'A{current_row}'], indent_level=1)
            apply_cell_style(pl_schedule_sheet[f'B{current_row}'])
            current_row += 1
        
        # Add total
        pl_schedule_sheet[f'A{current_row}'] = f"Total {category_name}"
        pl_schedule_sheet[f'B{current_row}'] = total
        
        apply_cell_style(pl_schedule_sheet[f'A{current_row}'], is_total=True)
        apply_cell_style(pl_schedule_sheet[f'B{current_row}'], is_total=True)
        current_row += 3
        
        schedule_num += 1
    
    # Notes Sheet
    notes_sheet = wb.create_sheet("Notes")
    
    notes_sheet['A1'] = "Notes to Financial Statements"
    notes_sheet['A1'].font = Font(bold=True, size=14)
    
    # Note 1: Capital Account
    notes_sheet['A3'] = "Note 1: Capital Account"
    notes_sheet['A3'].font = Font(bold=True, size=12)
    
    notes_sheet['A5'] = "Particulars"
    notes_sheet['B5'] = "Amount (₹)"
    
    apply_cell_style(notes_sheet['A5'], is_header=True)
    apply_cell_style(notes_sheet['B5'], is_header=True)
    
    notes_sheet['A6'] = "Opening Balance"
    notes_sheet['B6'] = st.session_state.financial_statements['notes']['note1_capital']['opening_balance']
    
    apply_cell_style(notes_sheet['A6'])
    apply_cell_style(notes_sheet['B6'])
    
    notes_sheet['A7'] = "Add: Capital Introduced"
    notes_sheet['B7'] = st.session_state.financial_statements['notes']['note1_capital']['additions']
    
    apply_cell_style(notes_sheet['A7'])
    apply_cell_style(notes_sheet['B7'])
    
    notes_sheet['A8'] = "Total"
    notes_sheet['B8'] = (
        st.session_state.financial_statements['notes']['note1_capital']['opening_balance'] +
        st.session_state.financial_statements['notes']['note1_capital']['additions']
    )
    
    apply_cell_style(notes_sheet['A8'], is_total=True)
    apply_cell_style(notes_sheet['B8'], is_total=True)
    
    # Set column widths
    for sheet in wb.worksheets:
        sheet.column_dimensions['A'].width = 40
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 20
    
    # Save the workbook to the BytesIO object
    wb.save(output)
    output.seek(0)
    
    return output

# Function to create a download link
def get_download_link(buffer, filename, text):
    b64 = base64.b64encode(buffer.getvalue()).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}">{text}</a>'
    return href

# Function to handle the "Add Sample Data" button
def add_sample_data():
    sample_data = [
        {'name': 'Land and Building', 'balance': 400000.00},
        {'name': 'Plant and Machinery', 'balance': 300000.00},
        {'name': 'Furniture and Fixtures', 'balance': 100000.00},
        {'name': 'Inventories', 'balance': 250000.00},
        {'name': 'Sundry Debtors', 'balance': 150000.00},
        {'name': 'Cash and Bank', 'balance': 200000.00},
        {'name': 'Capital Account', 'balance': 800000.00},
        {'name': 'Reserves', 'balance': 200000.00},
        {'name': 'Secured Loans', 'balance': 150000.00},
        {'name': 'Sundry Creditors', 'balance': 100000.00},
        {'name': 'Domestic Sales', 'balance': 1200000.00},
        {'name': 'Export Sales', 'balance': 300000.00},
        {'name': 'Interest Income', 'balance': 50000.00},
        {'name': 'Raw Material Consumed', 'balance': 600000.00},
        {'name': 'Salaries and Wages', 'balance': 300000.00},
        {'name': 'Interest Expenses', 'balance': 80000.00},
        {'name': 'Depreciation', 'balance': 120000.00},
        {'name': 'Administrative Expenses', 'balance': 100000.00},
        {'name': 'Selling Expenses', 'balance': 80000.00},
    ]
    
    st.session_state.tally_data = {
        'ledgers': sample_data,
        'tally_version': 'Sample Data',
        'export_date': datetime.now().isoformat()
    }
    
    # Pre-populate some mappings for the sample data
    st.session_state.mapped_accounts = {
        'Land and Building': 'BS_FixedAssets - Fixed Assets',
        'Plant and Machinery': 'BS_FixedAssets - Fixed Assets',
        'Furniture and Fixtures': 'BS_FixedAssets - Fixed Assets',
        'Inventories': 'BS_CurrentAssets - Current Assets',
        'Sundry Debtors': 'BS_CurrentAssets - Current Assets',
        'Cash and Bank': 'BS_CurrentAssets - Current Assets',
        'Capital Account': 'BS_Capital - Capital Account',
        'Reserves': 'BS_Reserves - Reserves & Surplus',
        'Secured Loans': 'BS_LongTermLoans - Long Term Loans',
        'Sundry Creditors': 'BS_CurrentLiabilities - Current Liabilities',
        'Domestic Sales': 'PL_Revenue - Revenue from Operations',
        'Export Sales': 'PL_Revenue - Revenue from Operations',
        'Interest Income': 'PL_OtherIncome - Other Income',
        'Raw Material Consumed': 'PL_COGS - Cost of Goods Sold',
        'Salaries and Wages': 'PL_EmployeeBenefits - Employee Benefits',
        'Interest Expenses': 'PL_FinanceCost - Finance Cost',
        'Depreciation': 'PL_Depreciation - Depreciation',
        'Administrative Expenses': 'PL_OtherExpenses - Other Expenses',
        'Selling Expenses': 'PL_OtherExpenses - Other Expenses',
    }
    
    # Pre-populate sub-schedule mappings
    st.session_state.sub_schedule_mapping = {
        'Land and Building': 'BS_FixedAssets_LandandBuildings - Land and Buildings',
        'Plant and Machinery': 'BS_FixedAssets_PlantandMachinery - Plant and Machinery',
        'Furniture and Fixtures': 'BS_FixedAssets_FurnitureandFixtures - Furniture and Fixtures',
        'Inventories': 'BS_CurrentAssets_Inventories - Inventories',
        'Sundry Debtors': 'BS_CurrentAssets_SundryDebtors - Sundry Debtors',
        'Cash and Bank': 'BS_CurrentAssets_CashandBankBalances - Cash and Bank Balances',
        'Capital Account': 'BS_CapitalAccount_OwnersCapital - Owner\'s Capital',
        'Domestic Sales': 'PL_RevenuefromOperations_DomesticSales - Domestic Sales',
        'Export Sales': 'PL_RevenuefromOperations_ExportSales - Export Sales',
        'Administrative Expenses': 'PL_OtherExpenses_AdministrativeExpenses - Administrative Expenses',
        'Selling Expenses': 'PL_OtherExpenses_SellingandDistributionExpenses - Selling and Distribution Expenses',
    }
    
    return st.session_state.tally_data

# Load saved mappings
main_mappings, sub_mappings = load_mappings()
st.session_state.mapped_accounts = main_mappings
st.session_state.sub_schedule_mapping = sub_mappings

# Main application header
st.title("Financial Statements Preparation System")
st.markdown("Prepare financial statements from Tally trial balance with multi-level tagging")

# Navigation
tabs = ["Upload Files", "Account Mapping", "View Statements", "Version History"]
selected_tab = st.sidebar.radio("Navigation", tabs)

# Sidebar for view statements options
if selected_tab == "View Statements":
    st.sidebar.markdown("### Select Statement")
    statement_type = st.sidebar.radio(
        "Statement Type",
        ["Balance Sheet", "Profit & Loss", "Notes", "Sub-Schedules"]
    )

# Debug option
debug_mode = st.sidebar.checkbox("Enable Debug Mode")
st.session_state.debug_mode = debug_mode

# Add "Use Sample Data" button in sidebar
if st.sidebar.button("Use Sample Data"):
    add_sample_data()
    st.sidebar.success("Sample data loaded!")
    rerun_app()

# Main content area based on selected tab
if selected_tab == "Upload Files":
    st.header("Upload Files")
    
    # File uploader for Tally XML/text file
    uploaded_file = st.file_uploader("Upload your Tally trial balance file", type=["xml", "txt"])
    
    # Add ability to paste Tally data
    st.markdown("### Or paste your Tally data below")
    pasted_data = st.text_area("Paste Tally trial balance data here", height=200)
    
    process_data = False
    data_source = None
    
    if uploaded_file is not None:
        # Process the uploaded file
        process_data = True
        data_source = uploaded_file.read()
        
    elif pasted_data:
        # Process the pasted data
        process_data = True
        data_source = pasted_data
    
    if process_data and data_source:
        try:
            # Show raw data for debugging
            if debug_mode:
                with st.expander("Raw Data Preview"):
                    if isinstance(data_source, bytes):
                        st.write("First 1000 characters:")
                        st.code(data_source[:1000].decode('utf-8', errors='replace'))
                    else:
                        st.write("First 1000 characters:")
                        st.code(data_source[:1000])
            
            # Parse the data
            parsed_data = parse_tally_file(data_source)
            st.session_state.tally_data = parsed_data
            
            # Check for new ledgers if we have previous mappings
            if st.session_state.mapped_accounts:
                new_ledgers = identify_new_ledgers(parsed_data['ledgers'], st.session_state.mapped_accounts)
                st.session_state.new_ledgers = new_ledgers
            
            # File information
            st.subheader("File Information")
            col1, col2 = st.columns(2)
            with col1:
                st.info(f"Tally Version: {parsed_data['tally_version']}")
            with col2:
                st.info(f"Export Date: {datetime.fromisoformat(parsed_data['export_date']).strftime('%Y-%m-%d %H:%M:%S')}")
            
            st.info(f"Ledgers Found: {len(parsed_data['ledgers'])}")
            
            # Show all ledgers in a dataframe
            st.subheader("Ledgers Found")
            ledger_df = pd.DataFrame(parsed_data['ledgers'])
            ledger_df['balance'] = ledger_df['balance'].apply(lambda x: f"₹{x:,.2f}")
            st.dataframe(ledger_df)
            
            # Button to proceed to mapping
            if st.button("Proceed to Mapping"):
                rerun_app()
                
        except Exception as e:
            st.error(f"Error processing data: {str(e)}")
            if debug_mode:
                st.exception(e)
                
                # Manual data entry option if parsing fails
                st.subheader("Manual Ledger Entry")
                st.write("Since automatic parsing failed, you can enter ledgers manually or use sample data.")
                
                # Use sample data button
                if st.button("Use Sample Data"):
                    add_sample_data()
                    st.success("Sample data loaded!")
                    rerun_app()
                
                # Create a manual entry section
                st.subheader("Or Enter Ledgers Manually")
                
                # Create a sample ledger list
                if 'manual_ledgers' not in st.session_state:
                    st.session_state.manual_ledgers = [
                        {'name': 'Capital Account', 'balance': 1000000},
                        {'name': 'Fixed Assets', 'balance': 800000},
                        {'name': 'Current Assets', 'balance': 700000}
                    ]
                
                # Display and edit the ledgers
                for i, ledger in enumerate(st.session_state.manual_ledgers):
                    cols = st.columns([3, 2])
                    with cols[0]:
                        st.session_state.manual_ledgers[i]['name'] = st.text_input(
                            f"Ledger Name {i+1}", 
                            value=ledger['name'],
                            key=f"manual_name_{i}"
                        )
                    with cols[1]:
                        st.session_state.manual_ledgers[i]['balance'] = st.number_input(
                            f"Balance {i+1}",
                            value=float(ledger['balance']),
                            key=f"manual_balance_{i}"
                        )
                
                # Add new ledger button
                if st.button("Add Another Ledger"):
                    st.session_state.manual_ledgers.append({'name': f'New Ledger {len(st.session_state.manual_ledgers)+1}', 'balance': 0})
                    rerun_app()
                
                # Use manual ledgers button
                if st.button("Use Manual Ledgers"):
                    st.session_state.tally_data = {
                        'ledgers': st.session_state.manual_ledgers,
                        'tally_version': 'Manual Entry',
                        'export_date': datetime.now().isoformat()
                    }
                    st.success("Manual ledgers added successfully!")
                    rerun_app()
    
    # Optional: Also allow uploading Excel template
    st.markdown("---")
    excel_template = st.file_uploader("Upload Financial Statement Excel Template (Optional)", type=["xlsx"])
    if excel_template is not None:
        try:
            # Read the Excel file
            excel_data = excel_template.read()
            st.session_state.excel_template = excel_data
            
            # Try to display information about the template
            wb = openpyxl.load_workbook(BytesIO(excel_data))
            sheet_names = wb.sheetnames
            
            st.success(f"Excel template uploaded successfully! Contains {len(sheet_names)} sheets.")
            
            if debug_mode:
                st.write("Sheet names:", ", ".join(sheet_names))
                
                # Show a preview of the first sheet
                first_sheet = wb[sheet_names[0]]
                data = []
                for row in first_sheet.iter_rows(max_row=5):
                    row_data = []
                    for cell in row:
                        row_data.append(cell.value)
                    data.append(row_data)
                
                st.write("Preview of first sheet:")
                st.dataframe(pd.DataFrame(data))
        except Exception as e:
            st.error(f"Error processing Excel template: {str(e)}")
            if debug_mode:
                st.exception(e)

elif selected_tab == "Account Mapping":
    st.header("Account Mapping with Sub-Schedules")
    
    # Check if we have data to work with
    if st.session_state.tally_data is None:
        st.warning("Please upload or paste Tally data first.")
        if st.button("Go to Upload"):
            rerun_app()
    else:
        # Add a search box for filtering ledgers
        search_term = st.text_input("Search ledgers", "")
        
        # Save mappings button
        if st.button("Save Mapping"):
            save_mappings()
        
        # Show warning for new ledgers
        if st.session_state.new_ledgers:
            st.warning(f"{len(st.session_state.new_ledgers)} new ledgers detected. Please map them below.")
            
            # Optional: Show the new ledgers separately
            new_ledger_names = [ledger['name'] for ledger in st.session_state.new_ledgers]
            with st.expander("View new ledgers"):
                st.write(", ".join(new_ledger_names))
        
        # Create mapping UI
        st.subheader("Map Ledgers to Financial Statement Items and Sub-Schedules")
        
        # Get the main category options
        mapping_options = create_mapping_options()
        
        # Filter ledgers based on search term
        if search_term:
            filtered_ledgers = [
                ledger for ledger in st.session_state.tally_data['ledgers'] 
                if search_term.lower() in ledger['name'].lower()
            ]
        else:
            filtered_ledgers = st.session_state.tally_data['ledgers']
        
        # Group the ledgers by mapping
        if st.checkbox("Group by mapping"):
            # Get unique mappings
            unique_mappings = set()
            for ledger in filtered_ledgers:
                mapping = st.session_state.mapped_accounts.get(ledger['name'], "Unmapped")
                unique_mappings.add(mapping)
            
            for mapping in sorted(unique_mappings):
                with st.expander(f"{mapping} ({sum(1 for ledger in filtered_ledgers if st.session_state.mapped_accounts.get(ledger['name'], 'Unmapped') == mapping)} ledgers)"):
                    # Create columns for a table-like display
                    for ledger in filtered_ledgers:
                        if st.session_state.mapped_accounts.get(ledger['name'], "Unmapped") == mapping:
                            # Create row with columns
                            cols = st.columns([2, 1.5, 3, 3])
                            
                            is_new = ledger['name'] in [l['name'] for l in st.session_state.new_ledgers]
                            prefix = "🆕 " if is_new else ""
                            
                            with cols[0]:
                                st.write(f"{prefix}{ledger['name']}")
                            with cols[1]:
                                st.write(f"₹{ledger['balance']:,.2f}")
                            
                            # Main category dropdown
                            with cols[2]:
                                # Create a unique key for each selectbox
                                index = next((i for i, opt in enumerate(mapping_options) if opt.startswith(mapping)), 0)
                                selected_mapping = st.selectbox(
                                    "Main Category",
                                    options=mapping_options,
                                    index=index,
                                    key=f"mapping_{ledger['name']}",
                                    label_visibility="collapsed"
                                )
                                
                                # Update mapping in session state when changed
                                if selected_mapping != "Select mapping...":
                                    st.session_state.mapped_accounts[ledger['name']] = selected_mapping
                            
                            # Sub-category dropdown
                            with cols[3]:
                                current_mapping = st.session_state.mapped_accounts.get(ledger['name'], "Select mapping...")
                                if current_mapping != "Select mapping...":
                                    # Get sub-categories for this main category
                                    sub_options = create_sub_category_options(current_mapping)
                                    
                                    # Get current sub-category selection
                                    current_sub = st.session_state.sub_schedule_mapping.get(ledger['name'], "Select sub-category...")
                                    sub_index = next((i for i, opt in enumerate(sub_options) if opt == current_sub), 0)
                                    
                                    selected_sub = st.selectbox(
                                        "Sub-Category",
                                        options=sub_options,
                                        index=sub_index,
                                        key=f"sub_mapping_{ledger['name']}",
                                        label_visibility="collapsed"
                                    )
                                    
                                    # Update sub-category mapping
                                    if selected_sub != "Select sub-category...":
                                        st.session_state.sub_schedule_mapping[ledger['name']] = selected_sub
                                else:
                                    st.write("Select main category first")
                            
                            # Add a subtle divider
                            st.markdown("---")
        else:
            # Show all ledgers together
            # Create a container with a scrollable area
            with st.container():
                # Create header for columns
                header_cols = st.columns([2, 1.5, 3, 3])
                with header_cols[0]:
                    st.write("**Ledger Name**")
                with header_cols[1]:
                    st.write("**Balance**")
                with header_cols[2]:
                    st.write("**Main Category**")
                with header_cols[3]:
                    st.write("**Sub-Schedule**")
                
                st.markdown("---")
                
                # Paginate the ledgers for better performance
                page_size = 15
                total_pages = (len(filtered_ledgers) + page_size - 1) // page_size
                
                if total_pages > 1:
                    page = st.number_input("Page", min_value=1, max_value=total_pages, value=1)
                    start_idx = (page - 1) * page_size
                    end_idx = min(start_idx + page_size, len(filtered_ledgers))
                    current_ledgers = filtered_ledgers[start_idx:end_idx]
                    st.write(f"Showing ledgers {start_idx+1}-{end_idx} of {len(filtered_ledgers)}")
                else:
                    current_ledgers = filtered_ledgers
                
                # Iterate through ledgers and create mapping inputs
                for ledger in current_ledgers:
                    ledger_name = ledger['name']
                    is_new = ledger_name in [l['name'] for l in st.session_state.new_ledgers]
                    
                    # Add a visual indicator for new ledgers
                    prefix = "🆕 " if is_new else ""
                    
                    # Get current mapping value
                    current_mapping = st.session_state.mapped_accounts.get(ledger_name, "Select mapping...")
                    
                    # Create row with columns
                    cols = st.columns([2, 1.5, 3, 3])
                    with cols[0]:
                        st.write(f"{prefix}{ledger_name}")
                    with cols[1]:
                        st.write(f"₹{ledger['balance']:,.2f}")
                    
                    # Main category dropdown
                    with cols[2]:
                        # Create a unique key for each selectbox
                        index = next((i for i, opt in enumerate(mapping_options) if opt == current_mapping), 0)
                        selected_mapping = st.selectbox(
                            "Main Category",
                            options=mapping_options,
                            index=index,
                            key=f"mapping_{ledger_name}",
                            label_visibility="collapsed"
                        )
                        
                        # Update mapping in session state when changed
                        if selected_mapping != "Select mapping...":
                            st.session_state.mapped_accounts[ledger_name] = selected_mapping
                    
                    # Sub-category dropdown
                    with cols[3]:
                        current_mapping = st.session_state.mapped_accounts.get(ledger_name, "Select mapping...")
                        if current_mapping != "Select mapping...":
                            # Get sub-categories for this main category
                            sub_options = create_sub_category_options(current_mapping)
                            
                            # Get current sub-category selection
                            current_sub = st.session_state.sub_schedule_mapping.get(ledger_name, "Select sub-category...")
                            sub_index = next((i for i, opt in enumerate(sub_options) if opt == current_sub), 0)
                            
                            selected_sub = st.selectbox(
                                "Sub-Category",
                                options=sub_options,
                                index=sub_index,
                                key=f"sub_mapping_{ledger_name}",
                                label_visibility="collapsed"
                            )
                            
                            # Update sub-category mapping
                            if selected_sub != "Select sub-category...":
                                st.session_state.sub_schedule_mapping[ledger_name] = selected_sub
                        else:
                            st.write("Select main category first")
                    
                    # Add a subtle divider
                    st.markdown("---")
        
        # Button to generate financial statements
        if st.button("Generate Financial Statements"):
            generate_financial_statements()
            st.success("Financial statements generated successfully!")
            rerun_app()

elif selected_tab == "View Statements":
    st.header("View Financial Statements")
    
    # Check if we have generated financial statements
    if st.session_state.financial_statements is None:
        st.warning("No financial statements have been generated yet.")
        if st.button("Go to Account Mapping"):
            rerun_app()
    else:
        # Show generation timestamp
        st.info(f"Generated on: {datetime.fromisoformat(st.session_state.financial_statements['generated_at']).strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Show selected statement type
        if statement_type == "Balance Sheet":
            st.subheader("Balance Sheet")
            
            # Create columns for better layout
            cols = st.columns([4, 1, 2])
            with cols[0]:
                st.markdown("#### Particulars")
            with cols[1]:
                st.markdown("#### Note No.")
            with cols[2]:
                st.markdown("#### Amount (₹)")
            
            # EQUITY AND LIABILITIES
            st.markdown("---")
            cols = st.columns([4, 1, 2])
            with cols[0]:
                st.markdown("**EQUITY AND LIABILITIES**")
            
            # Capital Account
            cols = st.columns([4, 1, 2])
            with cols[0]:
                st.markdown("&nbsp;&nbsp;&nbsp;&nbsp;Capital Account")
            with cols[1]:
                st.markdown("1")
            with cols[2]:
                st.markdown(f"₹{st.session_state.financial_statements['balance_sheet']['liabilities']['capital']['total']:,.2f}")
            
            # Reserves and Surplus
            cols = st.columns([4, 1, 2])
            with cols[0]:
                st.markdown("&nbsp;&nbsp;&nbsp;&nbsp;Reserves and Surplus")
            with cols[1]:
                st.markdown("2")
            with cols[2]:
                st.markdown(f"₹{st.session_state.financial_statements['balance_sheet']['liabilities']['reserves']['total']:,.2f}")
            
            # Long Term Loans
            cols = st.columns([4, 1, 2])
            with cols[0]:
                st.markdown("&nbsp;&nbsp;&nbsp;&nbsp;Long Term Loans")
            with cols[1]:
                st.markdown("3")
            with cols[2]:
                st.markdown(f"₹{st.session_state.financial_statements['balance_sheet']['liabilities']['long_term_loans']['total']:,.2f}")
            
            # Current Liabilities
            cols = st.columns([4, 1, 2])
            with cols[0]:
                st.markdown("&nbsp;&nbsp;&nbsp;&nbsp;Current Liabilities")
            with cols[1]:
                st.markdown("4")
            with cols[2]:
                st.markdown(f"₹{st.session_state.financial_statements['balance_sheet']['liabilities']['current_liabilities']['total']:,.2f}")
            
            # Total Liabilities
            st.markdown("---")
            cols = st.columns([4, 1, 2])
            with cols[0]:
                st.markdown("**Total Liabilities**")
            with cols[2]:
                total_liabilities = (
                    st.session_state.financial_statements['balance_sheet']['liabilities']['capital']['total'] +
                    st.session_state.financial_statements['balance_sheet']['liabilities']['reserves']['total'] +
                    st.session_state.financial_statements['balance_sheet']['liabilities']['long_term_loans']['total'] +
                    st.session_state.financial_statements['balance_sheet']['liabilities']['current_liabilities']['total']
                )
                st.markdown(f"**₹{total_liabilities:,.2f}**")
            
            # ASSETS
            st.markdown("---")
            cols = st.columns([4, 1, 2])
            with cols[0]:
                st.markdown("**ASSETS**")
            
            # Fixed Assets
            cols = st.columns([4, 1, 2])
            with cols[0]:
                st.markdown("&nbsp;&nbsp;&nbsp;&nbsp;Fixed Assets")
            with cols[1]:
                st.markdown("5")
            with cols[2]:
                st.markdown(f"₹{st.session_state.financial_statements['balance_sheet']['assets']['fixed_assets']['total']:,.2f}")
            
            # Investments
            cols = st.columns([4, 1, 2])
            with cols[0]:
                st.markdown("&nbsp;&nbsp;&nbsp;&nbsp;Investments")
            with cols[1]:
                st.markdown("6")
            with cols[2]:
                st.markdown(f"₹{st.session_state.financial_statements['balance_sheet']['assets']['investments']['total']:,.2f}")
            
            # Current Assets
            cols = st.columns([4, 1, 2])
            with cols[0]:
                st.markdown("&nbsp;&nbsp;&nbsp;&nbsp;Current Assets")
            with cols[1]:
                st.markdown("7")
            with cols[2]:
                st.markdown(f"₹{st.session_state.financial_statements['balance_sheet']['assets']['current_assets']['total']:,.2f}")
            
            # Total Assets
            st.markdown("---")
            cols = st.columns([4, 1, 2])
            with cols[0]:
                st.markdown("**Total Assets**")
            with cols[2]:
                total_assets = (
                    st.session_state.financial_statements['balance_sheet']['assets']['fixed_assets']['total'] +
                    st.session_state.financial_statements['balance_sheet']['assets']['investments']['total'] +
                    st.session_state.financial_statements['balance_sheet']['assets']['current_assets']['total']
                )
                st.markdown(f"**₹{total_assets:,.2f}**")
        
        elif statement_type == "Profit & Loss":
            st.subheader("Statement of Profit and Loss")
            
            # Create columns for better layout
            cols = st.columns([4, 1, 2])
            with cols[0]:
                st.markdown("#### Particulars")
            with cols[1]:
                st.markdown("#### Note No.")
            with cols[2]:
                st.markdown("#### Amount (₹)")
            
            # INCOME
            st.markdown("---")
            cols = st.columns([4, 1, 2])
            with cols[0]:
                st.markdown("**INCOME**")
            
            # Revenue
            cols = st.columns([4, 1, 2])
            with cols[0]:
                st.markdown("&nbsp;&nbsp;&nbsp;&nbsp;Revenue from Operations")
            with cols[1]:
                st.markdown("8")
            with cols[2]:
                st.markdown(f"₹{st.session_state.financial_statements['profit_and_loss']['income']['revenue']['total']:,.2f}")
            
            # Other Income
            cols = st.columns([4, 1, 2])
            with cols[0]:
                st.markdown("&nbsp;&nbsp;&nbsp;&nbsp;Other Income")
            with cols[1]:
                st.markdown("9")
            with cols[2]:
                st.markdown(f"₹{st.session_state.financial_statements['profit_and_loss']['income']['other_income']['total']:,.2f}")
            
            # Total Income
            st.markdown("---")
            cols = st.columns([4, 1, 2])
            with cols[0]:
                st.markdown("**Total Income**")
            with cols[2]:
                total_income = (
                    st.session_state.financial_statements['profit_and_loss']['income']['revenue']['total'] +
                    st.session_state.financial_statements['profit_and_loss']['income']['other_income']['total']
                )
                st.markdown(f"**₹{total_income:,.2f}**")
            
            # EXPENSES
            st.markdown("---")
            cols = st.columns([4, 1, 2])
            with cols[0]:
                st.markdown("**EXPENSES**")
            
            # COGS
            cols = st.columns([4, 1, 2])
            with cols[0]:
                st.markdown("&nbsp;&nbsp;&nbsp;&nbsp;Cost of Goods Sold")
            with cols[1]:
                st.markdown("10")
            with cols[2]:
                st.markdown(f"₹{st.session_state.financial_statements['profit_and_loss']['expenses']['cogs']['total']:,.2f}")
            
            # Employee Benefits
            cols = st.columns([4, 1, 2])
            with cols[0]:
                st.markdown("&nbsp;&nbsp;&nbsp;&nbsp;Employee Benefits Expense")
            with cols[1]:
                st.markdown("11")
            with cols[2]:
                st.markdown(f"₹{st.session_state.financial_statements['profit_and_loss']['expenses']['employee_benefits']['total']:,.2f}")
            
            # Finance Costs
            cols = st.columns([4, 1, 2])
            with cols[0]:
                st.markdown("&nbsp;&nbsp;&nbsp;&nbsp;Finance Costs")
            with cols[1]:
                st.markdown("12")
            with cols[2]:
                st.markdown(f"₹{st.session_state.financial_statements['profit_and_loss']['expenses']['finance_costs']['total']:,.2f}")
            
            # Depreciation
            cols = st.columns([4, 1, 2])
            with cols[0]:
                st.markdown("&nbsp;&nbsp;&nbsp;&nbsp;Depreciation")
            with cols[1]:
                st.markdown("13")
            with cols[2]:
                st.markdown(f"₹{st.session_state.financial_statements['profit_and_loss']['expenses']['depreciation']['total']:,.2f}")
            
            # Other Expenses
            cols = st.columns([4, 1, 2])
            with cols[0]:
                st.markdown("&nbsp;&nbsp;&nbsp;&nbsp;Other Expenses")
            with cols[1]:
                st.markdown("14")
            with cols[2]:
                st.markdown(f"₹{st.session_state.financial_statements['profit_and_loss']['expenses']['other_expenses']['total']:,.2f}")
            
            # Total Expenses
            st.markdown("---")
            cols = st.columns([4, 1, 2])
            with cols[0]:
                st.markdown("**Total Expenses**")
            with cols[2]:
                total_expenses = (
                    st.session_state.financial_statements['profit_and_loss']['expenses']['cogs']['total'] +
                    st.session_state.financial_statements['profit_and_loss']['expenses']['employee_benefits']['total'] +
                    st.session_state.financial_statements['profit_and_loss']['expenses']['finance_costs']['total'] +
                    st.session_state.financial_statements['profit_and_loss']['expenses']['depreciation']['total'] +
                    st.session_state.financial_statements['profit_and_loss']['expenses']['other_expenses']['total']
                )
                st.markdown(f"**₹{total_expenses:,.2f}**")
            
            # Profit Before Tax
            st.markdown("---")
            cols = st.columns([4, 1, 2])
            with cols[0]:
                st.markdown("**Profit Before Tax**")
            with cols[2]:
                profit = total_income - total_expenses
                st.markdown(f"**₹{profit:,.2f}**")
        
        elif statement_type == "Notes":
            st.subheader("Notes to Financial Statements")
            
            # Note 1: Capital Account
            st.markdown("### Note 1: Capital Account")
            
            # Create a DataFrame for better presentation
            note1_data = {
                'Particulars': ['Opening Balance', 'Add: Capital Introduced', 'Total'],
                'Amount (₹)': [
                    st.session_state.financial_statements['notes']['note1_capital']['opening_balance'],
                    st.session_state.financial_statements['notes']['note1_capital']['additions'],
                    st.session_state.financial_statements['notes']['note1_capital']['opening_balance'] + 
                    st.session_state.financial_statements['notes']['note1_capital']['additions']
                ]
            }
            note1_df = pd.DataFrame(note1_data)
            note1_df['Amount (₹)'] = note1_df['Amount (₹)'].apply(lambda x: f"₹{x:,.2f}")
            st.table(note1_df)
            
            # Other notes would be added here
        
        elif statement_type == "Sub-Schedules":
            st.subheader("Sub-Schedules")
            
            # Choose between BS and PL schedules
            schedule_type = st.radio("Select schedule type", ["Balance Sheet Schedules", "Profit & Loss Schedules"])
            
            if schedule_type == "Balance Sheet Schedules":
                # Show Balance Sheet schedules
                for category_key, category_data in st.session_state.financial_statements['sub_schedules'].items():
                    if not category_key.startswith('BS_'):
                        continue
                        
                    # Skip empty schedules
                    if not any(item['amount'] > 0 for item in category_data['items'].values()):
                        continue
                        
                    with st.expander(f"Schedule: {category_data['name']}"):
                        # Create table
                        data = []
                        total = 0
                        
                        for sub_key, sub_data in category_data['items'].items():
                            if sub_data['amount'] > 0:
                                data.append({
                                    'Particulars': sub_data['name'],
                                    'Amount (₹)': f"₹{sub_data['amount']:,.2f}"
                                })
                                total += sub_data['amount']
                        
                        # Add total row
                        data.append({
                            'Particulars': f"Total {category_data['name']}",
                            'Amount (₹)': f"₹{total:,.2f}"
                        })
                        
                        # Show the table
                        st.table(pd.DataFrame(data))
                        
                        # Show ledgers in this category if available
                        ledgers_in_category = []
                        for sub_key, sub_data in category_data['items'].items():
                            if 'ledgers' in sub_data and sub_data['ledgers']:
                                for ledger in sub_data['ledgers']:
                                    ledgers_in_category.append({
                                        'Ledger Name': ledger['name'],
                                        'Sub-Category': sub_data['name'],
                                        'Amount (₹)': f"₹{ledger['balance']:,.2f}"
                                    })
                        
                        if ledgers_in_category:
                            st.markdown("#### Mapped Ledgers")
                            st.dataframe(pd.DataFrame(ledgers_in_category))
            
            else:  # Profit & Loss Schedules
                # Show Profit & Loss schedules
                for category_key, category_data in st.session_state.financial_statements['sub_schedules'].items():
                    if not category_key.startswith('PL_'):
                        continue
                        
                    # Skip empty schedules
                    if not any(item['amount'] > 0 for item in category_data['items'].values()):
                        continue
                        
                    with st.expander(f"Schedule: {category_data['name']}"):
                        # Create table
                        data = []
                        total = 0
                        
                        for sub_key, sub_data in category_data['items'].items():
                            if sub_data['amount'] > 0:
                                data.append({
                                    'Particulars': sub_data['name'],
                                    'Amount (₹)': f"₹{sub_data['amount']:,.2f}"
                                })
                                total += sub_data['amount']
                        
                        # Add total row
                        data.append({
                            'Particulars': f"Total {category_data['name']}",
                            'Amount (₹)': f"₹{total:,.2f}"
                        })
                        
                        # Show the table
                        st.table(pd.DataFrame(data))
                        
                        # Show ledgers in this category if available
                        ledgers_in_category = []
                        for sub_key, sub_data in category_data['items'].items():
                            if 'ledgers' in sub_data and sub_data['ledgers']:
                                for ledger in sub_data['ledgers']:
                                    ledgers_in_category.append({
                                        'Ledger Name': ledger['name'],
                                        'Sub-Category': sub_data['name'],
                                        'Amount (₹)': f"₹{ledger['balance']:,.2f}"
                                    })
                        
                        if ledgers_in_category:
                            st.markdown("#### Mapped Ledgers")
                            st.dataframe(pd.DataFrame(ledgers_in_category))
        
        # Export options
        st.markdown("---")
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("Export as Excel"):
                excel_buffer = export_to_excel()
                st.markdown(
                    get_download_link(excel_buffer, "financial_statements.xlsx", "Download Excel File"),
                    unsafe_allow_html=True
                )
        
        with col2:
            if st.button("Export as PDF"):
                st.info("PDF export functionality would be implemented here")

elif selected_tab == "Version History":
    st.header("Version History")
    
    if not st.session_state.versions:
        st.warning("No version history available yet.")
    else:
        # Create a table of versions
        version_data = []
        for version in st.session_state.versions:
            version_data.append({
                'Version': f"Version {version['id']}",
                'Generated On': datetime.fromisoformat(version['timestamp']).strftime('%Y-%m-%d %H:%M:%S'),
                'Total Ledgers': len(version['tally_data']['ledgers']),
                'Current': "✓" if version['id'] == st.session_state.current_version else ""
            })
        
        version_df = pd.DataFrame(version_data)
        st.table(version_df)
        
        # Load a previous version
        st.subheader("Load a Previous Version")
        version_options = [f"Version {v['id']}" for v in st.session_state.versions]
        selected_version = st.selectbox("Select Version", version_options)
        
        if st.button("Load Selected Version"):
            version_id = int(selected_version.split(" ")[1])
            version = next((v for v in st.session_state.versions if v['id'] == version_id), None)
            
            if version:
                st.session_state.current_version = version['id']
                st.session_state.mapped_accounts = version['mapped_accounts'].copy()
                if 'sub_schedule_mapping' in version:
                    st.session_state.sub_schedule_mapping = version['sub_schedule_mapping'].copy()
                if 'financial_statements' in version:
                    st.session_state.financial_statements = version['financial_statements']
                st.success(f"Loaded {selected_version}")
                rerun_app()

# Add some styling
st.markdown("""
<style>
    .reportview-container {
        background-color: #f0f2f6
    }
    .sidebar .sidebar-content {
        background-color: #f9f9f9
    }
    h1 {
        color: #0f4c81
    }
    h2 {
        color: #0f4c81
    }
    h3 {
        color: #0f4c81
    }
</style>
""", unsafe_allow_html=True)